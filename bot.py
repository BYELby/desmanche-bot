import os

import discord
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

ARQUIVO = "desmanche.xlsx"
ID_SERVIDOR = 1506310709263208639
CANAL_REGISTRO_ID_1 = 1506362648621482004
CANAL_REGISTRO_ID_2 = 1506362648621482004
EQUIPES = []
ADM_ID = 401135654349832193
CARGO_REGISTRADO_ID = 1506376039763415222
CANAL_RANKING_ID = 1506388755551096892
MENSAGEM_RANKING_ID = None

async def atualizar_rankings():
    global MENSAGEM_RANKING_ID

    canal = bot.get_channel(CANAL_RANKING_ID)

    if canal is None:
        return

    wb = load_workbook(ARQUIVO)

    ranking_ws = wb["Ranking"]

    jogadores = []
    equipes = {}

    for row in range(2, ranking_ws.max_row + 1):

        nome = ranking_ws.cell(row, 1).value
        equipe = ranking_ws.cell(row, 2).value
        pontos = ranking_ws.cell(row, 3).value
        carros = ranking_ws.cell(row, 4).value

        if nome and pontos is not None:

            jogadores.append((nome, equipe, pontos, carros))

            if equipe not in equipes:
                equipes[equipe] = {
                    "pontos": 0,
                    "carros": 0
                }

            equipes[equipe]["pontos"] += pontos
            equipes[equipe]["carros"] += carros

    jogadores.sort(key=lambda x: x[2], reverse=True)

    ranking_equipes = sorted(
        equipes.items(),
        key=lambda x: x[1]["pontos"],
        reverse=True
    )

    texto = "🏆 **RANKING DESMANCHE** 🏆\n\n"

    texto += "## 👑 TOP JOGADORES\n\n"

    if jogadores:
        for i, jogador in enumerate(jogadores[:10], start=1):

            medalha = ""

            if i == 1:
                medalha = "🥇"
            elif i == 2:
                medalha = "🥈"
            elif i == 3:
                medalha = "🥉"

            nome, equipe, pontos, carros = jogador

            texto += (
                f"{medalha} **{i}º** {nome} "
                f"| {pontos} pts "
                f"| {carros} carros "
                f"| {equipe}\n"
            )

    else:
        texto += "Nenhum registro.\n"

    texto += "\n## 🏁 TOP EQUIPES\n\n"

    if ranking_equipes:

        for i, (equipe, dados) in enumerate(ranking_equipes[:10], start=1):

            medalha = ""

            if i == 1:
                medalha = "🥇"
            elif i == 2:
                medalha = "🥈"
            elif i == 3:
                medalha = "🥉"

            texto += (
                f"{medalha} **{i}º** {equipe} "
                f"| {dados['pontos']} pts "
                f"| {dados['carros']} carros\n"
            )

    else:
        texto += "Nenhuma equipe.\n"

    if MENSAGEM_RANKING_ID:

        try:
            mensagem = await canal.fetch_message(MENSAGEM_RANKING_ID)

            await mensagem.edit(content=texto)

            return

        except:
            MENSAGEM_RANKING_ID = None

    mensagem = await canal.send(texto)

    MENSAGEM_RANKING_ID = mensagem.id



CATEGORIAS_CONCE = {
    "0-5k": 0,
    "5k-20k": 1,
    "20k-100k": 2,
    "100k-250k": 3,
    "250k-500k": 4,
    "500k-1.25kk": 5,
    "1.25kk-6kk": 8
}

CATEGORIAS_DIMA = {
    "10k-25k": 2,
    "25k-35k": 3,
    "35k-45k": 5,
    "45k-60k": 10
}

def criar_planilha():
    if not os.path.exists(ARQUIVO):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        ws.append(["Data", "Discord", "Nome", "Equipe", "Tipo", "Categoria", "Pontos"])

        ranking = wb.create_sheet("Ranking")
        ranking.append(["Nome", "Equipe", "Total de Pontos", "Total de Carros"])

        wb.save(ARQUIVO)

def salvar_registro(ctx, nome, equipe, tipo, categoria, pontos):
    wb = load_workbook(ARQUIVO)
    registros = wb["Registros"]
    ranking = wb["Ranking"]

    registros.append([
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        ctx.author.name,
        nome,
        equipe,
        tipo,
        categoria,
        pontos
    ])

    if pontos > 0:
        encontrado = False

        for row in range(2, ranking.max_row + 1):
            if ranking.cell(row, 1).value == nome and ranking.cell(row, 2).value == equipe:
                ranking.cell(row, 3).value += pontos
                ranking.cell(row, 4).value += 1
                encontrado = True
                break

        if not encontrado:
            ranking.append([nome, equipe, pontos, 1])

    wb.save(ARQUIVO)

criar_planilha()

bot = discord.Bot(debug_guilds=[ID_SERVIDOR])

@bot.event
async def on_ready():
    print(f"Logado como {bot.user}")

@bot.slash_command(name="conce", description="Registrar veículo de concessionária")
async def conce(
    ctx,
    nome: str,
    equipe: discord.Option(str, autocomplete=discord.utils.basic_autocomplete(EQUIPES)),
    categoria: discord.Option(str, choices=list(CATEGORIAS_CONCE.keys()))
):
    if ctx.channel.id != CANAL_REGISTRO_ID_1:
        await ctx.respond(f"❌ Use esse comando no canal correto: <#{CANAL_REGISTRO_ID}>", ephemeral=True)
        return

    pontos = CATEGORIAS_CONCE[categoria]

    salvar_registro(ctx, nome, equipe, "Conce", categoria, pontos)
    await atualizar_rankings()

    if categoria == "0-5k":
        mensagem = "foi roubar carrinho de mercado? KKKKKKKK"
    else:
        mensagem = (
        f"✅ Registro CONCE salvo!\n"
        f"Nome: {nome}\n"
        f"Equipe: {equipe}\n"
        f"Categoria: {categoria}\n"
        f"Pontos: {pontos}"
    )

    await ctx.respond(mensagem)

@bot.slash_command(name="dima", description="Registrar veículo de dima")
async def dima(
    ctx,
    nome: str,
    equipe: discord.Option(str, autocomplete=discord.utils.basic_autocomplete(EQUIPES)),
    categoria: discord.Option(str, choices=list(CATEGORIAS_DIMA.keys()))
):
    
    if ctx.channel.id != CANAL_REGISTRO_ID_2:
        await ctx.respond(f"❌ Use esse comando no canal correto: <#{CANAL_REGISTRO_ID_2}>", ephemeral=True)
        return

    pontos = CATEGORIAS_DIMA[categoria]

    salvar_registro(ctx, nome, equipe, "Dima", categoria, pontos)
    await atualizar_rankings()
    await ctx.respond(
        f"✅ Registro DIMA salvo!\n"
        f"Nome: {nome}\n"
        f"Equipe: {equipe}\n"
        f"Categoria: {categoria}\n"
        f"Pontos: {pontos}"
    )

@bot.slash_command(name="ranking_equipes", description="Mostrar ranking das equipes")
async def equipes(ctx):
    wb = load_workbook(ARQUIVO)
    ranking_ws = wb["Ranking"]

    equipes = {}

    for row in range(2, ranking_ws.max_row + 1):
        equipe = ranking_ws.cell(row, 2).value
        pontos = ranking_ws.cell(row, 3).value
        carros = ranking_ws.cell(row, 4).value

        if equipe:
            if equipe not in equipes:
                equipes[equipe] = {"pontos": 0, "carros": 0}

            equipes[equipe]["pontos"] += pontos or 0
            equipes[equipe]["carros"] += carros or 0

    if not equipes:
        await ctx.respond("📊 Ainda não existe ranking por equipes.")
        return

    ranking_equipes = sorted(
        equipes.items(),
        key=lambda item: item[1]["pontos"],
        reverse=True
    )

    texto = "🏁 **RANKING POR EQUIPES** 🏁\n\n"

    for i, (equipe, dados) in enumerate(ranking_equipes, start=1):
        texto += f"**{i}º** - {equipe} | {dados['pontos']} pts | {dados['carros']} carros\n"

    await ctx.respond(texto)

@bot.slash_command(name="criar_equipe", description="Criar equipe")
async def criar_equipe(ctx, nome: str):

    if ctx.author.id != ADM_ID:
        await ctx.respond("❌ Você não tem permissão.", ephemeral=True)
        return

    if nome in EQUIPES:
        await ctx.respond("❌ Essa equipe já existe.")
        return

    EQUIPES.append(nome)

    await ctx.respond(f"✅ Equipe '{nome}' criada com sucesso!")

@bot.slash_command(name="resetar_registros", description="Resetar campeonato")
async def resetar_registros(ctx):

    if ctx.author.id != ADM_ID:
        await ctx.respond("❌ Você não tem permissão.", ephemeral=True)
        return

    wb = load_workbook(ARQUIVO)

    registros = wb["Registros"]
    ranking = wb["Ranking"]

    registros.delete_rows(2, registros.max_row)
    ranking.delete_rows(2, ranking.max_row)

    wb.save(ARQUIVO)

    await ctx.respond("⚠️ Todos os registros e rankings foram resetados.")

class RegistroModal(discord.ui.Modal):
    def __init__(self):
        super().__init__(title="Registro")

        self.nome = discord.ui.InputText(
            label="Nome do personagem e id",
            placeholder="Ex: Gabryel Sobek | #923",
            required=True
        )

        self.equipe = discord.ui.InputText(
            label="Equipe",
            placeholder="Ex: NKT",
            required=True
        )

        self.add_item(self.nome)
        self.add_item(self.equipe)

    async def callback(self, interaction: discord.Interaction):
        nome = self.nome.value
        equipe = self.equipe.value

        try:
            await interaction.user.edit(nick=nome)

            cargo = discord.utils.get(interaction.guild.roles, name=equipe)

            if cargo is None:
                cargo = await interaction.guild.create_role(name=equipe)

            await interaction.user.add_roles(cargo)
            cargo_registrado = interaction.guild.get_role(CARGO_REGISTRADO_ID)
            if cargo_registrado:
                await interaction.user.add_roles(cargo_registrado)

            await interaction.response.send_message(
                f"✅ Registro concluído!\n"
                f"Nome alterado para: **{nome}**\n"
                f"Equipe: **{equipe}**",
                ephemeral=True
            )

        except discord.Forbidden:
            await interaction.response.send_message(
                "❌ Não consegui alterar seu nome ou adicionar cargo. Verifique as permissões do bot.",
                ephemeral=True
            )


class RegistroView(discord.ui.View):
    @discord.ui.button(label="Fazer registro", style=discord.ButtonStyle.green)
    async def fazer_registro(self, button, interaction):
        await interaction.response.send_modal(RegistroModal())


@bot.slash_command(name="painel_registro", description="Enviar painel de registro")
async def painel_registro(ctx):

    if ctx.author.id != ADM_ID:
        await ctx.respond("❌ Você não tem permissão.", ephemeral=True)
        return

    await ctx.respond("✅ Painel enviado.", ephemeral=True)

    await ctx.channel.send(
        "📋 **Faça seu registro**\n\n"
        "Clique no botão abaixo e informe seu nome e equipe.",
        view=RegistroView()
    )






TOKEN = os.getenv("DISCORD_TOKEN") 
bot.run(TOKEN)

 